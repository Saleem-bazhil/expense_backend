import os, sys, django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')
django.setup()

import openpyxl
from io import BytesIO
from datetime import datetime
from expenses.serializers import ExpenseCreateSerializer

# Create a real Excel file matching user's format:
# S.No | Date | Category | Branch | Type | Amount | Remark | Person | Mode | Balance | Actions
wb = openpyxl.Workbook()
ws = wb.active
ws.append(['S.No', 'Date', 'Category', 'Branch', 'Type', 'Amount', 'Remark', 'Person', 'Mode', 'Balance', 'Actions'])
ws.append([1, '2026-05-09', 'Food', 'Chennai', 'Debit', 500, 'Lunch', 'Vaseem', 'Cash', 5000, ''])
ws.append([2, '09-05-2026', 'Travel', 'Vellore', 'Credit', 1000, 'Refund', 'Ali', 'GPay', 6000, ''])
ws.append([3, datetime(2026, 5, 8), 'Petrol', 'Hosur', 'Debit', 300, 'Fuel', 'Kumar', 'UPI', 5700, ''])

buf = BytesIO()
wb.save(buf)
buf.seek(0)

# Now simulate what import_expenses does
wb2 = openpyxl.load_workbook(buf, data_only=True)
ws2 = wb2.active
rows = list(ws2.iter_rows(values_only=True))

headers = [str(cell).strip().lower() if cell is not None else '' for cell in rows[0]]
print("HEADERS:", headers)

header_mapping = {
    'date': 'date', 'category': 'category', 'branch': 'branch',
    'credit': 'credited_amount', 'debit': 'debited_amount',
    'credit amount': 'credited_amount', 'credit_amount': 'credited_amount', 'credited_amount': 'credited_amount',
    'credit remark': 'credit_remark', 'credit_remark': 'credit_remark',
    'credit person': 'credit_person', 'credit_person': 'credit_person',
    'credit payment mode': 'credit_payment_mode', 'credit_payment_mode': 'credit_payment_mode',
    'debit amount': 'debited_amount', 'debit_amount': 'debited_amount', 'debited_amount': 'debited_amount',
    'debit remark': 'debit_remark', 'debit_remark': 'debit_remark',
    'debit person': 'debit_person', 'debit_person': 'debit_person',
    'debit payment mode': 'debit_payment_mode', 'debit_payment_mode': 'debit_payment_mode',
    'remark': 'remark', 'person': 'person', 'mode': 'mode', 'payment mode': 'mode',
    'amount': 'amount', 'type': 'type', 'expense type': 'type', 'transaction type': 'type',
}

print("\nMAPPED HEADERS:")
for h in headers:
    print(f"  '{h}' -> {header_mapping.get(h, 'NOT MAPPED')}")

import_data = []
for row_idx, row in enumerate(rows[1:], 2):
    if not any(cell is not None and str(cell).strip() != '' for cell in row):
        continue

    row_dict = {}
    for col_idx, cell in enumerate(row):
        if col_idx < len(headers):
            header = headers[col_idx]
            field = header_mapping.get(header)
            if field:
                if cell is None:
                    row_dict[field] = None
                elif field == 'date':
                    import datetime as dt
                    if isinstance(cell, (dt.datetime, dt.date)):
                        if isinstance(cell, dt.datetime):
                            row_dict[field] = cell.date().isoformat()
                        else:
                            row_dict[field] = cell.isoformat()
                    else:
                        try:
                            date_str = str(cell).strip().split(' ')[0]
                             parsed_date = datetime.strptime(date_str, '%Y-%m-%d')
                            row_dict[field] = parsed_date.date().isoformat()
                        except ValueError:
                            try:
                                date_str = str(cell).strip().split(' ')[0]
                                parsed_date = datetime.strptime(date_str, '%d-%m-%Y')
                                row_dict[field] = parsed_date.date().isoformat()
                            except ValueError:
                                try:
                                    date_str = str(cell).strip().split(' ')[0]
                                    parsed_date = datetime.strptime(date_str, '%d/%m/%Y')
                                    row_dict[field] = parsed_date.date().isoformat()
                                except ValueError:
                                    row_dict[field] = str(cell).strip()
                elif field in ['credited_amount', 'debited_amount']:
                    try:
                        cleaned_val = str(cell).replace(',', '').strip()
                        val = float(cleaned_val)
                        row_dict[field] = val if val > 0 else None
                    except (ValueError, TypeError):
                        row_dict[field] = None
                elif field == 'amount':
                    try:
                        cleaned_val = str(cell).replace(',', '').strip()
                        row_dict['amount'] = float(cleaned_val)
                    except (ValueError, TypeError):
                        row_dict['amount'] = None
                elif field == 'type':
                    row_dict['type'] = str(cell).strip().lower()
                else:
                    row_dict[field] = str(cell).strip()

    # Handle amount/type
    if row_dict.get('amount') is not None:
        amt = row_dict['amount']
        t = row_dict.get('type', 'debit')
        if 'credit' in t:
            row_dict['credited_amount'] = amt
            row_dict['debited_amount'] = None
        else:
            row_dict['debited_amount'] = amt
            row_dict['credited_amount'] = None

    if 'credited_amount' not in row_dict:
        row_dict['credited_amount'] = None
    if 'debited_amount' not in row_dict:
        row_dict['debited_amount'] = None

    is_credit = row_dict.get('credited_amount') is not None

    if 'remark' in row_dict:
        if is_credit:
            row_dict['credit_remark'] = row_dict['remark']
            row_dict['debit_remark'] = ''
        else:
            row_dict['debit_remark'] = row_dict['remark']
            row_dict['credit_remark'] = ''

    if 'person' in row_dict:
        if is_credit:
            row_dict['credit_person'] = row_dict['person']
            row_dict['debit_person'] = ''
        else:
            row_dict['debit_person'] = row_dict['person']
            row_dict['credit_person'] = ''

    if 'mode' in row_dict:
        if is_credit:
            row_dict['credit_payment_mode'] = row_dict['mode']
            row_dict['debit_payment_mode'] = ''
        else:
            row_dict['debit_payment_mode'] = row_dict['mode']
            row_dict['credit_payment_mode'] = ''

    if not row_dict.get('branch'):
        row_dict['branch'] = 'Main Branch'
    if not row_dict.get('category'):
        row_dict['category'] = 'Misc'

    # Clean extra keys
    valid_fields = {
        'date', 'category', 'branch',
        'credited_amount', 'credit_remark', 'credit_person', 'credit_payment_mode',
        'debited_amount', 'debit_remark', 'debit_person', 'debit_payment_mode',
    }
    clean_data = {k: v for k, v in row_dict.items() if k in valid_fields}

    print(f"\n=== ROW {row_idx} ===")
    print(f"Raw row_dict: {row_dict}")
    print(f"Clean data:   {clean_data}")

    serializer = ExpenseCreateSerializer(data=clean_data)
    valid = serializer.is_valid()
    print(f"Is valid: {valid}")
    if not valid:
        print(f"ERRORS: {serializer.errors}")
    else:
        print("WOULD SAVE SUCCESSFULLY!")
