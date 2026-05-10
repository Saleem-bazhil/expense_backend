import csv
from datetime import datetime
from decimal import Decimal
from io import BytesIO

from django.contrib.auth import authenticate
from django.db.models import Sum, Q, F, Window
from django.db.models.functions import Coalesce, TruncMonth
from django.http import HttpResponse
from rest_framework import viewsets, status
from rest_framework.authtoken.models import Token
from rest_framework.decorators import api_view, permission_classes, action
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.response import Response

from rest_framework.pagination import PageNumberPagination

from .models import Branch, Expense, PaymentModeBalance, BillingReminder
from .serializers import BranchSerializer, ExpenseSerializer, ExpenseCreateSerializer, PaymentModeBalanceSerializer, BillingReminderSerializer


class ExpensePagination(PageNumberPagination):
    """Pagination that exposes `page_size` in every response — lets the
    frontend compute total pages without hardcoding the size."""

    def get_paginated_response(self, data):
        return Response({
            'count': self.page.paginator.count,
            'next': self.get_next_link(),
            'previous': self.get_previous_link(),
            'page_size': self.get_page_size(self.request),
            'results': data,
        })


class BranchViewSet(viewsets.ModelViewSet):
    """CRUD for branches."""
    queryset = Branch.objects.all()
    serializer_class = BranchSerializer
    pagination_class = None


class ExpenseViewSet(viewsets.ModelViewSet):
    """CRUD for expenses with filtering and running balance."""

    pagination_class = ExpensePagination

    def get_serializer_class(self):
        if self.action in ('create', 'update', 'partial_update'):
            return ExpenseCreateSerializer
        return ExpenseSerializer

    def get_queryset(self):
        qs = Expense.objects.select_related('branch').all()

        # Filter by branch (can be ID or location name)
        branch_val = self.request.query_params.get('branch')
        if branch_val:
            if branch_val.isdigit():
                qs = qs.filter(branch_id=branch_val)
            else:
                qs = qs.filter(branch__location__icontains=branch_val)

        # Filter by category
        category = self.request.query_params.get('category')
        if category:
            qs = qs.filter(category=category)

        # Filter by date range
        date_from = self.request.query_params.get('date_from')
        date_to = self.request.query_params.get('date_to')
        if date_from:
            qs = qs.filter(date__gte=date_from)
        if date_to:
            qs = qs.filter(date__lte=date_to)

        # Search
        search = self.request.query_params.get('search')
        if search:
            qs = qs.filter(
                Q(credit_remark__icontains=search) |
                Q(debit_remark__icontains=search) |
                Q(credit_person__icontains=search) |
                Q(debit_person__icontains=search) |
                Q(category__icontains=search)
            )

        return qs.order_by('date', 'created_at')

    def list(self, request, *args, **kwargs):
        """Override list to include running balance computation."""
        queryset = self.get_queryset()

        # Calculate running balance
        expenses = list(queryset)
        initial_balances = {m.payment_mode: m.initial_balance for m in PaymentModeBalance.objects.all()}
        balances = {}
        for expense in expenses:
            credit = expense.credited_amount or Decimal('0.00')
            debit = expense.debited_amount or Decimal('0.00')
            
            if debit > 0:
                mode = expense.debit_payment_mode or 'Other'
                if mode not in balances:
                    balances[mode] = initial_balances.get(mode, Decimal('0.00'))
                balances[mode] -= debit
                
            if credit > 0:
                mode = expense.credit_payment_mode or 'Other'
                if mode not in balances:
                    balances[mode] = initial_balances.get(mode, Decimal('0.00'))
                balances[mode] += credit
                
            expense.running_balances = balances.copy()

        # To show newest first, reverse the list AFTER calculating running balances,
        # then paginate. This keeps correct balance cache on each object.
        expenses.reverse()

        # Pagination
        page = self.paginate_queryset(expenses)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = self.get_serializer(expenses, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['delete'], url_path='delete-all')
    def delete_all(self, request):
        """Utility endpoint to delete all expenses."""
        count, _ = Expense.objects.all().delete()
        return Response({'detail': f'Successfully deleted {count} expenses.'}, status=status.HTTP_200_OK)


@api_view(['GET'])
def dashboard_view(request):
    """Aggregated dashboard stats."""
    qs = Expense.objects.all()

    # Apply same filters
    branch_val = request.query_params.get('branch')
    if branch_val:
        if branch_val.isdigit():
            qs = qs.filter(branch_id=branch_val)
        else:
            qs = qs.filter(branch__location__icontains=branch_val)

    category = request.query_params.get('category')
    if category:
        qs = qs.filter(category=category)

    date_from = request.query_params.get('date_from')
    date_to = request.query_params.get('date_to')
    if date_from:
        qs = qs.filter(date__gte=date_from)
    if date_to:
        qs = qs.filter(date__lte=date_to)

    totals = qs.aggregate(
        total_credits=Coalesce(Sum('credited_amount'), Decimal('0.00')),
        total_debits=Coalesce(Sum('debited_amount'), Decimal('0.00')),
    )
    
    # Total Balance is the sum of all Payment Mode balances (Company-wide absolute balance)
    total_initial = PaymentModeBalance.objects.aggregate(t=Coalesce(Sum('initial_balance'), Decimal('0.00')))['t']
    global_credits = Expense.objects.aggregate(t=Coalesce(Sum('credited_amount'), Decimal('0.00')))['t']
    global_debits = Expense.objects.aggregate(t=Coalesce(Sum('debited_amount'), Decimal('0.00')))['t']
    total_balance = total_initial + global_credits - global_debits

    # Category breakdown (for pie chart)
    category_data = (
        qs.values('category')
        .annotate(
            total_credit=Coalesce(Sum('credited_amount'), Decimal('0.00')),
            total_debit=Coalesce(Sum('debited_amount'), Decimal('0.00')),
        )
        .order_by('category')
    )

    # Monthly trend (for line chart)
    monthly_data = (
        qs.annotate(month=TruncMonth('date'))
        .values('month')
        .annotate(
            credits=Coalesce(Sum('credited_amount'), Decimal('0.00')),
            debits=Coalesce(Sum('debited_amount'), Decimal('0.00')),
        )
        .order_by('month')
    )

    # Branch-wise (for bar chart and detailed breakdown)
    branch_qs = (
        qs.values('branch__location')
        .annotate(
            total_credit=Coalesce(Sum('credited_amount'), Decimal('0.00')),
            total_debit=Coalesce(Sum('debited_amount'), Decimal('0.00')),
        )
        .order_by('branch__location')
    )

    branch_breakdown = []
    for item in branch_qs:
        location = item['branch__location']
        # Category breakdown for THIS branch (scoped to current filters)
        branch_cats = (
            qs.filter(branch__location=location)
            .values('category')
            .annotate(
                total_credit=Coalesce(Sum('credited_amount'), Decimal('0.00')),
                total_debit=Coalesce(Sum('debited_amount'), Decimal('0.00')),
            )
            .order_by('-total_debit')
        )

        branch_breakdown.append({
            'branch': location,
            'total_credit': str(item['total_credit']),
            'total_debit': str(item['total_debit']),
            'category_breakdown': [
                {
                    'category': c['category'],
                    'total_credit': str(c['total_credit']),
                    'total_debit': str(c['total_debit']),
                }
                for c in branch_cats
            ]
        })

    return Response({
        'total_balance': str(total_balance),
        'total_credits': str(totals['total_credits']),
        'total_debits': str(totals['total_debits']),
        'category_breakdown': [
            {
                'category': item['category'],
                'total_credit': str(item['total_credit']),
                'total_debit': str(item['total_debit']),
            }
            for item in category_data
        ],
        'monthly_trend': [
            {
                'month': item['month'].strftime('%Y-%m') if item['month'] else '',
                'credits': str(item['credits']),
                'debits': str(item['debits']),
            }
            for item in monthly_data
        ],
        'branch_breakdown': branch_breakdown,
    })


@api_view(['GET'])
def export_expenses(request):
    """Export expenses to CSV."""
    qs = Expense.objects.select_related('branch').all().order_by('date', 'created_at')

    # Apply filters
    branch_val = request.query_params.get('branch')
    if branch_val:
        if branch_val.isdigit():
            qs = qs.filter(branch_id=branch_val)
        else:
            qs = qs.filter(branch__location__icontains=branch_val)

    category = request.query_params.get('category')
    if category:
        qs = qs.filter(category=category)

    date_from = request.query_params.get('date_from')
    date_to = request.query_params.get('date_to')
    if date_from:
        qs = qs.filter(date__gte=date_from)
    if date_to:
        qs = qs.filter(date__lte=date_to)

    # Note: avoid the name `format` — DRF reserves it for content negotiation.
    fmt = request.query_params.get('type', 'csv')

    if fmt == 'excel':
        # Excel export
        try:
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Expenses'

            headers = ['S.No', 'Date', 'Category', 'Branch', 'Credit Amount',
                        'Credit Remark', 'Credit Person', 'Credit Payment Mode',
                        'Debit Amount', 'Debit Remark', 'Debit Person',
                        'Debit Payment Mode', 'Running Balance']
            ws.append(headers)

            initial_balances = {m.payment_mode: m.initial_balance for m in PaymentModeBalance.objects.all()}
            balances = {}
            for idx, expense in enumerate(qs, 1):
                credit = expense.credited_amount or Decimal('0.00')
                debit = expense.debited_amount or Decimal('0.00')
                if debit > 0:
                    mode = expense.debit_payment_mode or 'Other'
                    if mode not in balances:
                        balances[mode] = initial_balances.get(mode, Decimal('0.00'))
                    balances[mode] -= debit
                    
                if credit > 0:
                    mode = expense.credit_payment_mode or 'Other'
                    if mode not in balances:
                        balances[mode] = initial_balances.get(mode, Decimal('0.00'))
                    balances[mode] += credit

                running_balance = " | ".join(f"{k}: {float(v)}" for k, v in balances.items() if v != Decimal('0.00'))


                ws.append([
                    idx,
                    expense.date.strftime('%Y-%m-%d'),
                    expense.category,
                    expense.branch.location,
                    float(credit),
                    expense.credit_remark,
                    expense.credit_person,
                    expense.credit_payment_mode,
                    float(debit),
                    expense.debit_remark,
                    expense.debit_person,
                    expense.debit_payment_mode,
                    running_balance,
                ])

            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            response = HttpResponse(
                buffer.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="expenses.xlsx"'
            return response
        except ImportError:
            return Response(
                {"error": "openpyxl not installed for Excel export"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    else:
        # CSV export
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="expenses.csv"'

        writer = csv.writer(response)
        writer.writerow(['S.No', 'Date', 'Category', 'Branch', 'Credit Amount',
                          'Credit Remark', 'Credit Person', 'Credit Payment Mode',
                          'Debit Amount', 'Debit Remark', 'Debit Person',
                          'Debit Payment Mode', 'Running Balance'])

        initial_balances = {m.payment_mode: m.initial_balance for m in PaymentModeBalance.objects.all()}
        balances = {}
        for idx, expense in enumerate(qs, 1):
            credit = expense.credited_amount or Decimal('0.00')
            debit = expense.debited_amount or Decimal('0.00')
            if debit > 0:
                mode = expense.debit_payment_mode or 'Other'
                if mode not in balances:
                    balances[mode] = initial_balances.get(mode, Decimal('0.00'))
                balances[mode] -= debit
                
            if credit > 0:
                mode = expense.credit_payment_mode or 'Other'
                if mode not in balances:
                    balances[mode] = initial_balances.get(mode, Decimal('0.00'))
                balances[mode] += credit

            running_balance = " | ".join(f"{k}: {float(v)}" for k, v in balances.items() if v != Decimal('0.00'))


            writer.writerow([
                idx,
                expense.date.strftime('%Y-%m-%d'),
                expense.category,
                expense.branch.location,
                credit,
                expense.credit_remark,
                expense.credit_person,
                expense.credit_payment_mode,
                debit,
                expense.debit_remark,
                expense.debit_person,
                expense.debit_payment_mode,
                running_balance,
            ])

        return response


# ---------------------------------------------------------------------------
# Payment Mode Balances
# ---------------------------------------------------------------------------
@api_view(['GET'])
def payment_mode_balances_view(request):
    """Return all payment modes with initial + current balance, including those only present in expenses.
    
    Supports filtering:
      - fy: Financial year, e.g. '2025-2026' (April to March)
      - date_from / date_to: Custom date range
    """
    # Build expense filter based on query params
    expense_filter = Q()
    fy = request.query_params.get('fy')
    month = request.query_params.get('month')
    date_from = request.query_params.get('date_from')
    date_to = request.query_params.get('date_to')

    start_date = None
    end_date = None

    if fy:
        try:
            parts = fy.split('-')
            start_year = int(parts[0])
            end_year = int(parts[1])
            
            if month:
                # month is 1-12
                m = int(month)
                # In India, FY starts in April. 
                # April (4) to Dec (12) are in start_year. 
                # Jan (1) to March (3) are in end_year.
                year = start_year if m >= 4 else end_year
                import calendar
                _, last_day = calendar.monthrange(year, m)
                start_date = f'{year}-{m:02d}-01'
                end_date = f'{year}-{m:02d}-{last_day:02d}'
            else:
                start_date = f'{start_year}-04-01'
                end_date = f'{end_year}-03-31'
            
            expense_filter &= Q(date__gte=start_date, date__lte=end_date)
        except (ValueError, IndexError):
            pass
    elif month:
        try:
            m = int(month)
            year = datetime.now().year
            import calendar
            _, last_day = calendar.monthrange(year, m)
            start_date = f'{year}-{m:02d}-01'
            end_date = f'{year}-{m:02d}-{last_day:02d}'
            expense_filter &= Q(date__gte=start_date, date__lte=end_date)
        except ValueError:
            pass
    else:
        if date_from:
            expense_filter &= Q(date__gte=date_from)
            start_date = date_from
        if date_to:
            expense_filter &= Q(date__lte=date_to)
            end_date = date_to

    balances = list(PaymentModeBalance.objects.all())
    explicit_modes = {b.payment_mode for b in balances}
    
    credit_modes = Expense.objects.exclude(credit_payment_mode='').values_list('credit_payment_mode', flat=True).distinct()
    debit_modes = Expense.objects.exclude(debit_payment_mode='').values_list('debit_payment_mode', flat=True).distinct()
    
    all_used_modes = set(credit_modes).union(set(debit_modes))
    missing_modes = all_used_modes - explicit_modes
    
    for mode in missing_modes:
        if mode:
            balances.append(PaymentModeBalance(
                id=len(balances) + 9999, # Fake ID to bypass frontend unique key warnings
                payment_mode=mode,
                initial_balance=Decimal('0.00')
            ))

    result = []
    for bal in balances:
        mode = bal.payment_mode
        
        # Calculate actual initial balance taking into account transactions before start_date
        period_initial = bal.initial_balance
        if start_date:
            past_credits = Expense.objects.filter(
                date__lt=start_date, credit_payment_mode=mode
            ).aggregate(
                total=Coalesce(Sum('credited_amount'), Decimal('0.00'))
            )['total']
            past_debits = Expense.objects.filter(
                date__lt=start_date, debit_payment_mode=mode
            ).aggregate(
                total=Coalesce(Sum('debited_amount'), Decimal('0.00'))
            )['total']
            period_initial += past_credits - past_debits

        # Credits with this payment mode (filtered)
        total_credits = Expense.objects.filter(
            expense_filter, credit_payment_mode=mode
        ).aggregate(
            total=Coalesce(Sum('credited_amount'), Decimal('0.00'))
        )['total']
        # Debits with this payment mode (filtered)
        total_debits = Expense.objects.filter(
            expense_filter, debit_payment_mode=mode
        ).aggregate(
            total=Coalesce(Sum('debited_amount'), Decimal('0.00'))
        )['total']
        
        current = period_initial + total_credits - total_debits
        period_available = total_credits - total_debits

        bal.initial_balance = period_initial
        bal.current_balance = current
        bal.total_credits = total_credits
        bal.total_debits = total_debits
        bal.period_available = period_available
        
        result.append(bal)

    serializer = PaymentModeBalanceSerializer(result, many=True)
    return Response(serializer.data)



@api_view(['POST'])
def payment_mode_balance_set(request):
    """Create or update initial balance for a payment mode."""
    mode = request.data.get('payment_mode', '').strip()
    initial = request.data.get('initial_balance')

    if not mode:
        return Response(
            {'detail': 'payment_mode is required.'},
            status=status.HTTP_400_BAD_REQUEST,
        )

    obj, created = PaymentModeBalance.objects.update_or_create(
        payment_mode=mode,
        defaults={'initial_balance': Decimal(str(initial or 0))},
    )

    # Compute current balance
    total_credits = Expense.objects.filter(
        credit_payment_mode=mode
    ).aggregate(
        total=Coalesce(Sum('credited_amount'), Decimal('0.00'))
    )['total']
    total_debits = Expense.objects.filter(
        debit_payment_mode=mode
    ).aggregate(
        total=Coalesce(Sum('debited_amount'), Decimal('0.00'))
    )['total']
    obj.current_balance = obj.initial_balance + total_credits - total_debits

    serializer = PaymentModeBalanceSerializer(obj)
    return Response(serializer.data, status=status.HTTP_200_OK)


@api_view(['DELETE'])
def payment_mode_balance_delete(request):
    """Delete a payment mode balance entry."""
    mode = request.data.get('payment_mode', '').strip()

    if not mode:
        return Response(
            {'detail': 'payment_mode is required.'},
            status=status.HTTP_400_BAD_REQUEST,
        )

    try:
        obj = PaymentModeBalance.objects.get(payment_mode=mode)
        obj.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)
    except PaymentModeBalance.DoesNotExist:
        return Response(
            {'detail': f'Payment mode "{mode}" not found.'},
            status=status.HTTP_404_NOT_FOUND,
        )


# ---------------------------------------------------------------------------
# Categories — expose model choices so the frontend doesn't hardcode them.
# ---------------------------------------------------------------------------
@api_view(['GET'])
def categories_view(request):
    """Single source of truth for expense categories (drawn from the model)."""
    return Response([value for value, _ in Expense.CATEGORY_CHOICES])


# ---------------------------------------------------------------------------
# Auth
# ---------------------------------------------------------------------------
@api_view(['POST'])
@permission_classes([AllowAny])
def login_view(request):
    """Username + password → token. Used by the SPA login form."""
    username = (request.data.get('username') or '').strip()
    password = request.data.get('password') or ''

    if not username or not password:
        return Response(
            {'detail': 'Username and password are required.'},
            status=status.HTTP_400_BAD_REQUEST,
        )

    user = authenticate(request, username=username, password=password)
    if user is None or not user.is_active:
        return Response(
            {'detail': 'Invalid credentials.'},
            status=status.HTTP_401_UNAUTHORIZED,
        )

    token, _ = Token.objects.get_or_create(user=user)
    return Response({
        'token': token.key,
        'username': user.username,
        'is_staff': user.is_staff,
    })


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def logout_view(request):
    """Invalidate the caller's token."""
    Token.objects.filter(user=request.user).delete()
    return Response(status=status.HTTP_204_NO_CONTENT)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def me_view(request):
    """Return current user info — used to verify a stored token is still valid."""
    return Response({
        'username': request.user.username,
        'is_staff': request.user.is_staff,
    })


# ---------------------------------------------------------------------------
# Billing Reminders
# ---------------------------------------------------------------------------
@api_view(['GET'])
def billing_reminders_list(request):
    """List all billing reminders."""
    reminders = BillingReminder.objects.all()
    serializer = BillingReminderSerializer(reminders, many=True)
    return Response(serializer.data)


@api_view(['POST'])
def billing_reminder_create(request):
    """Create a new billing reminder."""
    serializer = BillingReminderSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['PUT'])
def billing_reminder_update(request, pk):
    """Update a billing reminder."""
    try:
        reminder = BillingReminder.objects.get(pk=pk)
    except BillingReminder.DoesNotExist:
        return Response({'detail': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)

    serializer = BillingReminderSerializer(reminder, data=request.data, partial=True)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['PATCH'])
def billing_reminder_toggle_paid(request, pk):
    """Toggle the is_paid status of a billing reminder."""
    try:
        reminder = BillingReminder.objects.get(pk=pk)
    except BillingReminder.DoesNotExist:
        return Response({'detail': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)

    reminder.is_paid = not reminder.is_paid
    reminder.save()
    serializer = BillingReminderSerializer(reminder)
    return Response(serializer.data)


@api_view(['DELETE'])
def billing_reminder_delete(request, pk):
    """Delete a billing reminder."""
    try:
        reminder = BillingReminder.objects.get(pk=pk)
    except BillingReminder.DoesNotExist:
        return Response({'detail': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)

    reminder.delete()
    return Response(status=status.HTTP_204_NO_CONTENT)


@api_view(['POST'])
def import_expenses(request):
    """Import expenses from Excel or CSV file."""
    file_obj = request.FILES.get('file')
    if not file_obj:
        return Response({'detail': 'No file was uploaded.'}, status=status.HTTP_400_BAD_REQUEST)

    import openpyxl
    from io import BytesIO

    try:
        fname = file_obj.name.lower()
        if fname.endswith('.xlsx'):
            wb = openpyxl.load_workbook(file_obj, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
        elif fname.endswith('.csv') or fname.endswith('.txt'):
            import csv
            file_data = file_obj.read().decode('utf-8-sig').splitlines()
            reader = csv.reader(file_data)
            rows = list(reader)
        else:
            return Response({'detail': 'Only .xlsx and .csv file formats are supported.'}, status=status.HTTP_400_BAD_REQUEST)
    except Exception as e:
        return Response({'detail': f'Error reading file: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)

    if not rows:
        return Response({'detail': 'The uploaded file is empty.'}, status=status.HTTP_400_BAD_REQUEST)

    # Dynamically find the header row by looking for known keywords
    header_keywords = {'date', 'category', 'branch', 'amount', 'credit', 'debit', 'type', 'remark'}
    header_idx = 0
    headers = []

    for i, row in enumerate(rows):
        curr_headers = [str(cell).strip().lower() if cell is not None else '' for cell in row]
        # If this row has multiple header keywords, it's probably our header row
        if any(k in curr_headers or any(k in h for h in curr_headers) for k in header_keywords):
            headers = curr_headers
            header_idx = i
            break
    else:
        # Fallback to first row if keywords aren't matched explicitly
        headers = [str(cell).strip().lower() if cell is not None else '' for cell in rows[0]]
        header_idx = 0

    print(f"[IMPORT DEBUG] Found headers at row {header_idx + 1}: {headers}")
    print(f"[IMPORT DEBUG] Total data rows to process: {len(rows) - header_idx - 1}")
    
    header_mapping = {
        'date': 'date',
        'category': 'category',
        'branch': 'branch',
        'credit': 'credited_amount',
        'debit': 'debited_amount',
        'credit amount': 'credited_amount',
        'credit_amount': 'credited_amount',
        'credited_amount': 'credited_amount',
        'credit remark': 'credit_remark',
        'credit_remark': 'credit_remark',
        'credit person': 'credit_person',
        'credit_person': 'credit_person',
        'credit payment mode': 'credit_payment_mode',
        'credit_payment_mode': 'credit_payment_mode',
        'debit amount': 'debited_amount',
        'debit_amount': 'debited_amount',
        'debited_amount': 'debited_amount',
        'debit remark': 'debit_remark',
        'debit_remark': 'debit_remark',
        'debit person': 'debit_person',
        'debit_person': 'debit_person',
        'debit payment mode': 'debit_payment_mode',
        'debit_payment_mode': 'debit_payment_mode',
        'remark': 'remark',
        'person': 'person',
        'mode': 'mode',
        'payment mode': 'mode',
        'amount': 'amount',
        'type': 'type',
        'expense type': 'type',
        'transaction type': 'type',
    }
 
    import_data = []
    # Start processing data rows occurring AFTER the header row
    for row_idx, row in enumerate(rows[header_idx + 1:], header_idx + 2):
        if not any(cell is not None and str(cell).strip() != '' for cell in row):
            continue
 
        row_dict = {}
        for col_idx, cell in enumerate(row):
            if col_idx < len(headers):
                header = headers[col_idx]
                field = header_mapping.get(header)
                if field:
                    if cell is None or str(cell).strip() == '':
                        if field in ['credited_amount', 'debited_amount', 'amount']:
                            row_dict[field] = None
                        else:
                            row_dict[field] = ''
                    elif field == 'date':
                        import datetime as dt
                        if isinstance(cell, (dt.datetime, dt.date)):
                            if isinstance(cell, dt.datetime):
                                row_dict[field] = cell.date().isoformat()
                            else:
                                row_dict[field] = cell.isoformat()
                        else:
                            try:
                                date_str = str(cell).strip().split(' ')[0]
                                parsed_date = datetime.strptime(date_str, '%Y-%m-%d')
                                row_dict[field] = parsed_date.date().isoformat()
                            except ValueError:
                                try:
                                    date_str = str(cell).strip().split(' ')[0]
                                    parsed_date = datetime.strptime(date_str, '%d-%m-%Y')
                                    row_dict[field] = parsed_date.date().isoformat()
                                except ValueError:
                                    try:
                                        date_str = str(cell).strip().split(' ')[0]
                                        parsed_date = datetime.strptime(date_str, '%d/%m/%Y')
                                        row_dict[field] = parsed_date.date().isoformat()
                                    except ValueError:
                                        try:
                                            date_str = str(cell).strip().split(' ')[0]
                                            parsed_date = datetime.strptime(date_str, '%Y/%m/%d')
                                            row_dict[field] = parsed_date.date().isoformat()
                                        except ValueError:
                                            try:
                                                date_str = str(cell).strip().split(' ')[0]
                                                parsed_date = datetime.strptime(date_str, '%d.%m.%Y')
                                                row_dict[field] = parsed_date.date().isoformat()
                                            except ValueError:
                                                try:
                                                    date_str = str(cell).strip().split(' ')[0]
                                                    parsed_date = datetime.strptime(date_str, '%Y.%m.%d')
                                                    row_dict[field] = parsed_date.date().isoformat()
                                                except ValueError:
                                                    try:
                                                        date_str = str(cell).strip().split(' ')[0]
                                                        parsed_date = datetime.strptime(date_str, '%d-%b-%Y')
                                                        row_dict[field] = parsed_date.date().isoformat()
                                                    except ValueError:
                                                        try:
                                                            date_str = " ".join(str(cell).strip().split(' ')[:3])
                                                            parsed_date = datetime.strptime(date_str, '%d %b %Y')
                                                            row_dict[field] = parsed_date.date().isoformat()
                                                        except ValueError:
                                                            row_dict[field] = str(cell).strip()
                    elif field in ['credited_amount', 'debited_amount']:
                        try:
                            cleaned_val = str(cell).replace('₹', '').replace('Rs', '').replace(',', '').strip()
                            val = float(cleaned_val)
                            row_dict[field] = val if val > 0 else None
                        except (ValueError, TypeError):
                            row_dict[field] = None
                    elif field == 'amount':
                        try:
                            cleaned_val = str(cell).replace('₹', '').replace('Rs', '').replace(',', '').strip()
                            row_dict['amount'] = float(cleaned_val)
                        except (ValueError, TypeError):
                            row_dict['amount'] = None
                    elif field == 'type':
                        row_dict['type'] = str(cell).strip().lower()
                    else:
                        row_dict[field] = str(cell).strip()

        # Handle single 'amount' and 'type' column format if separate columns aren't filled
        if row_dict.get('amount') is not None:
            amt = row_dict['amount']
            t = row_dict.get('type', 'debit')
            if 'credit' in t:
                row_dict['credited_amount'] = amt
                row_dict['debited_amount'] = None
            else:
                row_dict['debited_amount'] = amt
                row_dict['credited_amount'] = None

        if 'credited_amount' not in row_dict:
            row_dict['credited_amount'] = None
        if 'debited_amount' not in row_dict:
            row_dict['debited_amount'] = None

        is_credit = row_dict.get('credited_amount') is not None

        if 'remark' in row_dict:
            if is_credit:
                row_dict['credit_remark'] = row_dict['remark']
                row_dict['debit_remark'] = ''
            else:
                row_dict['debit_remark'] = row_dict['remark']
                row_dict['credit_remark'] = ''

        if 'person' in row_dict:
            if is_credit:
                row_dict['credit_person'] = row_dict['person']
                row_dict['debit_person'] = ''
            else:
                row_dict['debit_person'] = row_dict['person']
                row_dict['credit_person'] = ''

        if 'mode' in row_dict:
            if is_credit:
                row_dict['credit_payment_mode'] = row_dict['mode']
                row_dict['debit_payment_mode'] = ''
            else:
                row_dict['debit_payment_mode'] = row_dict['mode']
                row_dict['credit_payment_mode'] = ''

        # Normalize payment modes to valid Title Case/Uppercase choice values
        def normalize_mode(val):
            if not val:
                return ''
            val_lower = str(val).strip().lower()
            if val_lower in ['cash', 'c']:
                return 'Cash'
            if val_lower in ['bank transfer', 'bank_transfer', 'bank', 'transfer']:
                return 'Bank Transfer'
            if val_lower in ['gpay', 'g-pay', 'google pay']:
                return 'GPay'
            if val_lower in ['phonepe', 'phone-pe', 'phone pe']:
                return 'PhonePe'
            if val_lower in ['upi']:
                return 'UPI'
            if val_lower in ['cheque']:
                return 'Cheque'
            return 'Other'

        if row_dict.get('credit_payment_mode'):
            row_dict['credit_payment_mode'] = normalize_mode(row_dict['credit_payment_mode'])
        if row_dict.get('debit_payment_mode'):
            row_dict['debit_payment_mode'] = normalize_mode(row_dict['debit_payment_mode'])

        if 'branch' not in row_dict or not row_dict['branch']:
            row_dict['branch'] = 'Main Branch'
        if 'category' not in row_dict or not row_dict['category']:
            row_dict['category'] = 'Misc'
        if 'date' not in row_dict or not row_dict['date']:
            import datetime as dt
            row_dict['date'] = dt.date.today().isoformat()

        # Final sanitization: convert any leftover None/missing for string fields to empty string
        string_fields = [
            'credit_remark', 'debit_remark',
            'credit_person', 'debit_person',
            'credit_payment_mode', 'debit_payment_mode',
            'category', 'branch'
        ]
        for f in string_fields:
            if row_dict.get(f) is None:
                row_dict[f] = ''

        import_data.append((row_idx, row_dict))

    errors = []
    success_count = 0

    # Fields accepted by the serializer
    valid_fields = {
        'date', 'category', 'branch',
        'credited_amount', 'credit_remark', 'credit_person', 'credit_payment_mode',
        'debited_amount', 'debit_remark', 'debit_person', 'debit_payment_mode',
    }

    for row_idx, data in import_data:
        # Strip out extra keys that aren't in the serializer
        clean_data = {k: v for k, v in data.items() if k in valid_fields}
        print(f"[IMPORT DEBUG] Row {row_idx} clean_data: {clean_data}")
        serializer = ExpenseCreateSerializer(data=clean_data)
        if serializer.is_valid():
            serializer.save()
            success_count += 1
            print(f"[IMPORT DEBUG] Row {row_idx}: SAVED OK")
        else:
            err_msg = ", ".join([f"{k}: {', '.join(v)}" for k, v in serializer.errors.items()])
            errors.append(f"Row {row_idx}: {err_msg}")
            print(f"[IMPORT DEBUG] Row {row_idx} ERRORS: {serializer.errors}")

    if errors:
        return Response({
            'detail': f'Import completed. Successfully imported {success_count} of {success_count + len(errors)} entries.',
            'errors': errors,
            'success_count': success_count
        }, status=status.HTTP_200_OK)

    return Response({
        'detail': f'Successfully imported {success_count} expenses.',
        'success_count': success_count
    }, status=status.HTTP_201_CREATED)
